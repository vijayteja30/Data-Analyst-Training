import openpyxl
import os, sys

"""
Permissions:
    r-4
    w-2
    x-1

Users:
    1. Root User
    2. Normal User
    3. Guest User

Documentation:
    1. https://www.datacamp.com/tutorial/python-excel-tutorial

Tasks:
    1. Create dummy data for csv and excel processing
    2. Create New rows and columns in Both Excel file and also CSV file as well
    3. Create New worksheet and rename it and add data to it.
    4. Save the sheet data and save excel file
    5. Read the data from the excel file and write it to the console with Dictionary format
    6. Read the data from the excel file and write it to the console with List
    7. Read the data from the excel file and write it to the console with Tuple format
    8. Read the created excel file and also csv files which were created above with the help of Pandas library and do manipulations like 
         filtering, sorting, grouping, etc.
    9. Know about Parquet files and write this csv data to parquet file with compression and without compression
    10. Read the parquet file and write it to the console with Dictionary, List, Tuple format
"""

# my_wb = openpyxl.Workbook()
# my_wb.create_sheet("XYZ")

# my_wb.create_sheet("ABC")
# my_sheet = my_wb.active
# my_sheet.title = "SiriShirisha"
# print("My sheet title: " + my_sheet.title)
# my_sheet["A1"] = "Name"
# my_sheet["B1"] = "Age"
# my_sheet.cell(row=2, column=1, value="Alice")
# my_sheet.cell(row=2, column=2, value=30)


# sheet2 = my_wb["XYZ"]
# sheet2["A1"] = "Hello"
# sheet2["B1"] = "World"
# sheet2
# sheet2.cell(row=2, column=1, value="Python")
# sheet2.cell(row=2, column=2, value="Excel")
# my_wb.save("Temp_Excel_creation.xlsx")
# print("Successfully updated XYZ")





# Load the workbook
# my_wb = openpyxl.Workbook()
# ws = my_wb.active
# ws.title = "Employees"

# headers = ["Number", "Name", "Age", "Gender", "Department", "Mobile number"]
# ws.append(headers)

# # dummy data (Name, Age, Gender, Department, Mobile)
# rows = [
#     ("Alice", 30, "F", "Finance", "+1-555-0101"),
#     ("Bob", 28, "M", "IT", "+1-555-0102"),
#     ("Carol", 35, "F", "HR", "+1-555-0103"),
#     ("Dave", 40, "M", "Operations", "+1-555-0104"),
#     ("Eve", 26, "F", "Marketing", "+1-555-0105"),
# ]

# # write rows with Number column
# for idx, r in enumerate(rows, start=1):
#     ws.append((idx, *r))

# my_wb.save("Temp_Excel_creation.xlsx")
# print("Saved Temp_Excel_creation.xlsx with headers and dummy data")







import openpyxl

# read the workbook we just created
wb = openpyxl.load_workbook("Temp_Excel_creation.xlsx", data_only=True)
ws = wb["Employees"] if "Employees" in wb.sheetnames else wb.active

rows = list(ws.iter_rows(values_only=True))
if not rows:
    print("No data found in sheet.")
else:
    headers = rows[0]
    data_rows = rows[1:]

    # print as dictionaries (header -> value)
    # new_rows = []
    for row in data_rows:
        # if row['product_category'] in ['electronics', 'home appliances', 'shoes']:
        #     new_rows.append(row)
        record = dict(zip(headers, row))
        print(record)

wb.close()